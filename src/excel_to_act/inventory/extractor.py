"""Core inventory extraction from openpyxl with source locations."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell

from excel_to_act.schemas import (
    CellInventory,
    CellKind,
    CoverageSummary,
    RangeInventory,
    SheetInventory,
    SourceLocation,
    WorkbookInventory,
    WorkbookManifest,
)


def _loc(path: Path, object_type: str, sheet: str | None = None, index: int | None = None, address: str | None = None, object_id: str | None = None) -> SourceLocation:
    return SourceLocation(workbook_path=str(path), sheet_name=sheet, sheet_index=index, address=address, object_type=object_type, object_id=object_id)


def _safe_value(value: Any) -> str | int | float | bool | None:
    if value is None or isinstance(value, str | int | float | bool):
        return value
    return str(value)


class OpenpyxlInventoryExtractor:
    name = "openpyxl_inventory"

    def extract(self, workbook_path: Path, manifest: WorkbookManifest) -> WorkbookInventory:
        workbook_path = workbook_path.expanduser().resolve()
        wb = load_workbook(workbook_path, data_only=False, read_only=False, keep_vba=workbook_path.suffix.lower() == ".xlsm")
        sheets: list[SheetInventory] = []
        workbook_ranges: list[RangeInventory] = []
        unsupported = list(manifest.unsupported_features)
        recognized = 0
        try:
            for i, ws in enumerate(wb.worksheets):
                sheet = SheetInventory(
                    source_location=_loc(workbook_path, "sheet", ws.title, i),
                    name=ws.title,
                    index=i,
                    max_row=ws.max_row or 0,
                    max_column=ws.max_column or 0,
                    state=ws.sheet_state,
                )
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value is None:
                            continue
                        assert isinstance(cell, Cell)
                        is_formula = isinstance(cell.value, str) and cell.value.startswith("=")
                        sheet.cells.append(
                            CellInventory(
                                source_location=_loc(workbook_path, "cell", ws.title, i, cell.coordinate, f"{ws.title}!{cell.coordinate}"),
                                address=cell.coordinate,
                                row=cell.row,
                                column=cell.column,
                                kind=CellKind.formula if is_formula else CellKind.literal,
                                value=None if is_formula else _safe_value(cell.value),
                                formula=str(cell.value) if is_formula else None,
                                data_type=cell.data_type,
                                number_format=cell.number_format,
                                style_id=getattr(cell, "style_id", None),
                            )
                        )
                        recognized += 1
                for merged in ws.merged_cells.ranges:
                    sheet.ranges.append(RangeInventory(source_location=_loc(workbook_path, "merged_range", ws.title, i, str(merged), str(merged)), address=str(merged), kind="merged_range"))
                    recognized += 1
                for table in ws.tables.values():
                    sheet.ranges.append(RangeInventory(source_location=_loc(workbook_path, "table", ws.title, i, table.ref, table.name), name=table.name, address=table.ref, kind="table", metadata={"display_name": table.displayName}))
                    recognized += 1
                self._layout(ws, workbook_path, i, sheet)
                recognized += len(sheet.layout_objects)
                sheets.append(sheet)

            for name, defined_name in wb.defined_names.items():
                try:
                    destinations = list(defined_name.destinations)
                except Exception:
                    destinations = []
                if not destinations:
                    loc = _loc(workbook_path, "defined_name", object_id=name)
                    workbook_ranges.append(RangeInventory(source_location=loc, name=name, address=str(getattr(defined_name, "attr_text", name)), kind="defined_name", metadata={"scope": "workbook"}))
                    recognized += 1
                for sheet_name, address in destinations:
                    idx = wb.sheetnames.index(sheet_name) if sheet_name in wb.sheetnames else None
                    workbook_ranges.append(RangeInventory(source_location=_loc(workbook_path, "defined_name", sheet_name, idx, address, name), name=name, address=address, kind="defined_name", metadata={"scope": "workbook"}))
                    recognized += 1
        finally:
            wb.close()
        opaque_count = len(unsupported)
        coverage = CoverageSummary(recognized_inventory_objects=recognized, unsupported_or_opaque_objects=opaque_count, discovered_workbook_objects=recognized + opaque_count)
        return WorkbookInventory(workbook_sha256=manifest.sha256, sheets=sheets, workbook_ranges=workbook_ranges, unsupported_features=unsupported, coverage=coverage)

    def _layout(self, ws: Any, workbook_path: Path, index: int, sheet: SheetInventory) -> None:
        for row_idx, dim in ws.row_dimensions.items():
            if dim.hidden or dim.height:
                sheet.layout_objects.append(RangeInventory(source_location=_loc(workbook_path, "row_layout", ws.title, index, str(row_idx), f"row:{row_idx}"), address=str(row_idx), kind="row_layout", metadata={"hidden": bool(dim.hidden), "height": dim.height}))
        for col, dim in ws.column_dimensions.items():
            if dim.hidden or dim.width:
                sheet.layout_objects.append(RangeInventory(source_location=_loc(workbook_path, "column_layout", ws.title, index, col, f"col:{col}"), address=col, kind="column_layout", metadata={"hidden": bool(dim.hidden), "width": dim.width}))
        if ws.freeze_panes:
            sheet.layout_objects.append(RangeInventory(source_location=_loc(workbook_path, "freeze_panes", ws.title, index, str(ws.freeze_panes), "freeze_panes"), address=str(ws.freeze_panes), kind="freeze_panes"))
        for dv in getattr(ws.data_validations, "dataValidation", []):
            sheet.layout_objects.append(RangeInventory(source_location=_loc(workbook_path, "data_validation", ws.title, index, str(dv.sqref), f"dv:{dv.sqref}"), address=str(dv.sqref), kind="data_validation", metadata={"type": dv.type, "formula1": dv.formula1, "formula2": dv.formula2}))
        for cf_range in getattr(ws.conditional_formatting, "_cf_rules", {}):
            sheet.layout_objects.append(RangeInventory(source_location=_loc(workbook_path, "conditional_formatting", ws.title, index, str(cf_range), f"cf:{cf_range}"), address=str(cf_range), kind="conditional_formatting"))
        for row in ws.iter_rows():
            for cell in row:
                if cell.comment:
                    sheet.layout_objects.append(RangeInventory(source_location=_loc(workbook_path, "comment", ws.title, index, cell.coordinate, f"comment:{cell.coordinate}"), address=cell.coordinate, kind="comment", metadata={"text": cell.comment.text, "author": cell.comment.author}))
                if cell.hyperlink:
                    sheet.layout_objects.append(RangeInventory(source_location=_loc(workbook_path, "hyperlink", ws.title, index, cell.coordinate, f"hyperlink:{cell.coordinate}"), address=cell.coordinate, kind="hyperlink", metadata={"target": cell.hyperlink.target, "location": cell.hyperlink.location}))
        if ws.protection and ws.protection.sheet:
            sheet.layout_objects.append(RangeInventory(source_location=_loc(workbook_path, "sheet_protection", ws.title, index, object_id="sheet_protection"), address="sheet", kind="sheet_protection"))
