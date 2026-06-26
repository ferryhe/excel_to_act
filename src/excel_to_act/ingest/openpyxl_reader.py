"""openpyxl-backed workbook manifest reader."""

from __future__ import annotations

import hashlib
from pathlib import Path

from openpyxl import load_workbook

from excel_to_act.ingest.ooxml_package import scan_ooxml_package
from excel_to_act.schemas import SheetManifest, WorkbookManifest


def workbook_sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


class OpenpyxlWorkbookReader:
    name = "openpyxl"

    def read_manifest(self, workbook_path: Path) -> WorkbookManifest:
        workbook_path = workbook_path.expanduser().resolve()
        parts, unsupported = scan_ooxml_package(workbook_path)
        if workbook_path.suffix.lower() not in {".xlsx", ".xlsm"}:
            return WorkbookManifest(
                workbook_path=str(workbook_path),
                file_name=workbook_path.name,
                file_size=workbook_path.stat().st_size,
                sha256=workbook_sha256(workbook_path),
                sheets=[],
                unsupported_features=unsupported,
            )
        wb = load_workbook(workbook_path, data_only=False, read_only=False, keep_vba=workbook_path.suffix.lower() == ".xlsm")
        try:
            sheets = [
                SheetManifest(
                    name=ws.title,
                    index=i,
                    state=ws.sheet_state,
                    max_row=ws.max_row or 0,
                    max_column=ws.max_column or 0,
                )
                for i, ws in enumerate(wb.worksheets)
            ]
            defined_names = getattr(wb, "defined_names", None)
            named_count = len(list(getattr(defined_names, "items", lambda: [])())) if defined_names is not None else 0
            calc = getattr(wb, "calculation", None)
            calc_mode = getattr(calc, "calcMode", None) or getattr(calc, "mode", None)
            return WorkbookManifest(
                workbook_path=str(workbook_path),
                file_name=workbook_path.name,
                file_size=workbook_path.stat().st_size,
                sha256=workbook_sha256(workbook_path),
                is_macro_enabled=workbook_path.suffix.lower() == ".xlsm" or any("vbaProject.bin" in p.name for p in parts),
                sheets=sheets,
                named_ranges_count=named_count,
                calc_mode=calc_mode,
                package_parts=parts,
                unsupported_features=unsupported,
            )
        finally:
            wb.close()
