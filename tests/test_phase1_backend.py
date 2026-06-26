from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from typer.testing import CliRunner

from excel_to_act.confirm.templates import ConfirmationTemplateBuilder
from excel_to_act.graph.builder import RegexFormulaGraphBuilder
from excel_to_act.ingest.openpyxl_reader import OpenpyxlWorkbookReader
from excel_to_act.inventory.extractor import OpenpyxlInventoryExtractor
from excel_to_act.interfaces.cli import app
from excel_to_act.plugins.registry import PluginRegistry
from excel_to_act.schemas import SourceLocation, WorkbookInventory
from excel_to_act.store.local_store import LocalArtifactStore


def make_fixture(path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Inputs"
    ws["A1"] = "Assumption rate"
    ws["B1"] = 0.05
    ws["B1"].number_format = "0.00%"
    ws["A2"] = "Premium"
    ws["B2"] = 100
    ws["C2"] = "=B1*B2"
    ws["D2"] = "=Outputs!B2+C2"
    ws["E2"] = "=SUM(B1:B2)"
    ws["A3"].comment = Comment("review me", "tester")
    ws["A4"] = "link"
    ws["A4"].hyperlink = "https://example.com"
    ws.merge_cells("A5:B5")
    ws.freeze_panes = "B2"
    ws.row_dimensions[3].hidden = True
    ws.column_dimensions["D"].hidden = True
    dv = DataValidation(type="decimal", operator="between", formula1="0", formula2="1")
    ws.add_data_validation(dv)
    dv.add(ws["B1"])
    ws.conditional_formatting.add("B1:B2", __import__("openpyxl").formatting.rule.CellIsRule(operator="greaterThan", formula=["0"]))
    ws["A7"] = "Field"
    ws["B7"] = "Value"
    ws["A8"] = "Premium"
    ws["B8"] = 100
    tab = __import__("openpyxl").worksheet.table.Table(displayName="InputTable", ref="A7:B8")
    ws.add_table(tab)
    out = wb.create_sheet("Outputs")
    out["A1"] = "Result"
    out["B2"] = "=Inputs!C2"
    hidden = wb.create_sheet("HiddenData")
    hidden.sheet_state = "hidden"
    hidden["A1"] = 42
    wb.defined_names.add(DefinedName("PremiumCell", attr_text="'Inputs'!$B$2"))
    wb.save(path)
    wb.close()
    return path


def test_cli_help() -> None:
    result = CliRunner().invoke(app, ["--help"])
    assert result.exit_code == 0
    assert "Inspect a workbook and produce Phase 1 artifacts" in result.output


def test_schema_roundtrip_and_source_validation(tmp_path: Path) -> None:
    fixture = make_fixture(tmp_path / "fixture.xlsx")
    manifest = OpenpyxlWorkbookReader().read_manifest(fixture)
    dumped = manifest.model_dump_json()
    assert manifest.model_validate_json(dumped).schema_version == "phase1.v1"
    with pytest.raises(ValueError):
        SourceLocation(object_type="cell", address="A1")


def test_unsupported_file_type_returns_typed_diagnostic(tmp_path: Path) -> None:
    bad = tmp_path / "legacy.xls"
    bad.write_bytes(b"not an OOXML workbook")
    manifest = OpenpyxlWorkbookReader().read_manifest(bad)
    assert manifest.sheets == []
    assert manifest.unsupported_features[0].feature_type == "file_type"
    assert manifest.unsupported_features[0].severity == "error"


def test_plugin_registry_dispatch() -> None:
    reg = PluginRegistry()
    plugin = object()
    reg.register("reader", "fake", plugin)
    assert reg.get("reader", "fake") is plugin
    assert reg.names("reader") == ["fake"]
    with pytest.raises(ValueError):
        reg.register("reader", "fake", object())


def test_manifest_inventory_graph_classification_confirmation(tmp_path: Path) -> None:
    fixture = make_fixture(tmp_path / "fixture.xlsx")
    manifest = OpenpyxlWorkbookReader().read_manifest(fixture)
    assert [s.name for s in manifest.sheets] == ["Inputs", "Outputs", "HiddenData"]
    assert any(s.state == "hidden" for s in manifest.sheets)
    inventory = OpenpyxlInventoryExtractor().extract(fixture, manifest)
    assert isinstance(WorkbookInventory.model_validate_json(inventory.model_dump_json()), WorkbookInventory)
    assert inventory.coverage.recognized_inventory_objects + inventory.coverage.unsupported_or_opaque_objects == inventory.coverage.discovered_workbook_objects
    inputs = next(s for s in inventory.sheets if s.name == "Inputs")
    assert all(c.source_location for c in inputs.cells)
    assert any(r.kind == "table" for r in inputs.ranges)
    assert any(o.kind == "data_validation" for o in inputs.layout_objects)
    assert any(o.kind == "conditional_formatting" for o in inputs.layout_objects)
    assert any(o.kind == "comment" for o in inputs.layout_objects)
    graph = RegexFormulaGraphBuilder().build(inventory)
    edge_labels = {(e.source, e.target) for e in graph.edges}
    assert any("Outputs!B2" in tgt for _, tgt in edge_labels)
    assert any("Inputs!B1:B2" in tgt for _, tgt in edge_labels)
    from excel_to_act.classify.classifier import RuleBasedClassifier

    classification = RuleBasedClassifier().classify(inventory, graph)
    assert len(classification.items) >= inventory.coverage.recognized_inventory_objects
    assert all(item.reasons for item in classification.items)
    template = ConfirmationTemplateBuilder().build(classification)
    assert template.schema_version == "phase1.v1"


def test_cli_inspect_writes_artifacts(tmp_path: Path) -> None:
    fixture = make_fixture(tmp_path / "fixture.xlsx")
    out = tmp_path / "artifacts"
    result = CliRunner().invoke(app, ["inspect", str(fixture), "--out", str(out)])
    assert result.exit_code == 0, result.output
    expected = [
        "workbook_manifest.json",
        "inventory.json",
        "dependency_graph.json",
        "module_classification.json",
        "confirmation_template.json",
        "run_metadata.json",
        "artifact_index.json",
    ]
    for name in expected:
        path = out / name
        assert path.exists(), name
    assert json.loads((out / "inventory.json").read_text())["schema_version"] == "phase1.v1"
    metadata = json.loads((out / "run_metadata.json").read_text())
    run_dir = out / "workbooks" / metadata["workbook_sha256"] / metadata["run_id"]
    assert run_dir.exists()
    assert (run_dir / "inventory.json").exists()
    assert LocalArtifactStore(out).read_run(metadata["workbook_sha256"], metadata["run_id"]).run_id == metadata["run_id"]

    second = CliRunner().invoke(app, ["inspect", str(fixture), "--out", str(out)])
    assert second.exit_code == 0, second.output
    index = json.loads((out / "workbooks" / metadata["workbook_sha256"] / "artifact_index.json").read_text())
    assert len(index["runs"]) == 2
    assert {run["run_id"] for run in index["runs"]} >= {metadata["run_id"]}
